import json
import math
from collections import Counter
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def strict_round(x, digits=2):
    multiplier = 10 ** digits
    sign = 1 if x >= 0 else -1
    return math.floor(x * multiplier + 0.5 * sign) / multiplier


TASK_NAME = 'i10'

INPUT_FILE = f'instances/{TASK_NAME}.json'
BASELINE_FILE = f'instances/baseline_{TASK_NAME}.json'
OUTPUT_FILE = f'instances/output_{TASK_NAME}.json'

REF_LABEL = f'{TASK_NAME} BASELINE'
OUR_LABEL = f'{TASK_NAME} OUR'


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_coords(point_id, depot, orders_by_id):
    if point_id == 0:
        return depot['x'], depot['y']
    order = orders_by_id[point_id]
    return order['x'], order['y']


def euclidean(p1, p2):
    return strict_round(math.hypot(p1[0] - p2[0], p1[1] - p2[1]), 2)


def _find_distance(x1, y1, x2, y2):
    return strict_round(math.hypot(x1 - x2, y1 - y2), 2)


def _build_arrival_times(vehicles, by_id):
    arrival = {}
    for v in vehicles:
        order_ids = [pid for pid in v['route'] if pid != 0]
        times = v['time']
        for oid, t in zip(order_ids, times):
            arrival[oid] = t
    return arrival


def _compute_loader_shift_time(loaders, vehicles, by_id, loader_speed):
    arrival = _build_arrival_times(vehicles, by_id)
    total = 0.0
    for ld in loaders:
        route = ld['route']
        if not route:
            continue
        first = by_id[route[0]]
        home_x, home_y = first['x'], first['y']
        shift_start = arrival[route[0]]
        time = shift_start + first['loader_service_time']
        for i in range(1, len(route)):
            prev = by_id[route[i - 1]]
            cur = by_id[route[i]]
            d = _find_distance(prev['x'], prev['y'], cur['x'], cur['y'])
            time += d / loader_speed
            time = max(time, arrival[route[i]])
            time += cur['loader_service_time']
        last = by_id[route[-1]]
        back = (
            _find_distance(last['x'], last['y'], home_x, home_y)
            / loader_speed
        )
        shift_end = time + back
        total += (shift_end - shift_start)
    return round(total, 2)


def _count_trips(route):
    """Число рейсов машины: [0, ..., 0] (ровно 2 нуля) - 1 рейс.
    [0, ..., 0, ..., 0] (N нулей, N>2) - N-1 рейсов (multi-trip:
    машина возвращалась в депо и уезжала снова)."""
    zeros = route.count(0)
    return max(1, zeros - 1)


def calc_cost(input_data, output_data):
    w = input_data['weights']
    depot = input_data['depot']
    orders = input_data['orders']
    vehicle_speed = input_data['vehicle_speed']
    by_id = {o['id']: o for o in orders}
    vehicles = output_data['vehicles']
    loaders = output_data.get('loaders', [])

    n_v, n_l = len(vehicles), len(loaders)

    total_dist = 0.0
    total_travel_time = 0.0
    for v in vehicles:
        r = v['route']
        for a, b in zip(r[:-1], r[1:]):
            leg_dist = euclidean(get_coords(a, depot, by_id),
                                 get_coords(b, depot, by_id))
            total_dist += leg_dist
            total_travel_time += strict_round(leg_dist / vehicle_speed, 2)

    loader_work_time = _compute_loader_shift_time(
        loaders, vehicles, by_id, input_data.get('loader_speed', 1)
    )

    visited = {pid for v in vehicles for pid in v['route'] if pid != 0}
    mand = {o['id'] for o in orders if not o.get('optional', 0)}
    opt = {o['id'] for o in orders if o.get('optional', 0)}
    missed_opt = sorted(opt - visited)
    missed_mand = sorted(mand - visited)

    cost = {
        'vehicles':  n_v * w['vehicle_salary'],
        'loaders':   n_l * w['loader_salary'],
        'fuel':      strict_round(total_travel_time * w['fuel_cost'], 2),
        'loader_w':  loader_work_time * w['loader_work'],
        'penalty':   len(missed_opt) * w['optional_order_penalty'],
    }
    cost['total'] = strict_round(sum(cost.values()), 2)

    orders_per_vehicle = [
        len([pid for pid in v['route'] if pid != 0]) for v in vehicles
    ]
    route_lens = Counter(orders_per_vehicle)
    chain_lens = Counter(len(ld['route']) for ld in loaders)

    trips_per_vehicle = [_count_trips(v['route']) for v in vehicles]
    trips_dist = Counter(trips_per_vehicle)
    multi_trip_vehicles = sum(1 for t in trips_per_vehicle if t > 1)

    cap = input_data['vehicle_capacity']
    util = [
        sum(by_id[oid]['volume'] for oid in v['route'] if oid != 0) / cap
        for v in vehicles
    ] if vehicles else [0]

    return {
        'n_vehicles': n_v,
        'n_loaders': n_l,
        'total_dist': round(total_dist, 2),
        'total_travel_time': round(total_travel_time, 2),
        'loader_work_time': loader_work_time,
        'cost': cost,
        'missed_optional': missed_opt,
        'missed_mandatory': missed_mand,
        'served_mandatory': len(mand & visited),
        'total_mandatory': len(mand),
        'served_optional': len(opt & visited),
        'total_optional': len(opt),
        'route_lens': dict(sorted(route_lens.items())),
        'chain_lens': dict(sorted(chain_lens.items())),
        'avg_util': round(sum(util) / len(util), 4),
        'avg_orders_per_vehicle': round(
            sum(orders_per_vehicle) / max(n_v, 1), 2),
        'avg_slots_per_loader': round(
            sum(len(ld['route']) for ld in loaders) / max(n_l, 1), 2),
        'multi_trip_vehicles': multi_trip_vehicles,
        'trips_dist': dict(sorted(trips_dist.items())),
    }


def analyze():
    input_data = load_json(INPUT_FILE)
    results = {}

    try:
        results[REF_LABEL] = calc_cost(input_data, load_json(BASELINE_FILE))
    except FileNotFoundError:
        pass

    try:
        results[OUR_LABEL] = calc_cost(input_data, load_json(OUTPUT_FILE))
    except FileNotFoundError:
        pass

    return input_data, results


def print_results(results):
    if not results:
        print("нет решений")
        return

    for label, r in results.items():
        c = r['cost']
        print(f"\n--- {label} ---")
        print(f"  машины:   {r['n_vehicles']:>3}    {c['vehicles']:>10.2f}")
        print(f"  грузчики: {r['n_loaders']:>3}    {c['loaders']:>10.2f}")
        print(
            f"  топливо (dist={r['total_dist']}, "
            f"travel_time={r['total_travel_time']}): {c['fuel']:>10.2f}"
        )
        print(f"  работа грузч (t={r['loader_work_time']}): "
              f"{c['loader_w']:>10.2f}")
        print(f"  штраф опц ({len(r['missed_optional'])} шт): "
              f"{c['penalty']:>10.2f}")
        print(f"  ИТОГО:    {c['total']:>10.2f}")
        print(f"  покрытие: обяз "
              f"{r['served_mandatory']}/{r['total_mandatory']}, "
              f"опц {r['served_optional']}/{r['total_optional']}")
        if r['missed_mandatory']:
            print(f"  !!! ПРОПУЩЕНЫ ОБЯЗАТЕЛЬНЫЕ: {r['missed_mandatory']}")
        print(f"  длины маршрутов: {r['route_lens']}")
        print(f"  длины цепочек:   {r['chain_lens']}")
        print(
            f"  multi-trip машин: {r['multi_trip_vehicles']} / "
            f"{r['n_vehicles']} "
            f"(распределение рейсов: {r['trips_dist']})"
        )

    if REF_LABEL in results and OUR_LABEL in results:
        ref = results[REF_LABEL]
        r = results[OUR_LABEL]
        dt = r['cost']['total'] - ref['cost']['total']
        base = ref['cost']['total']
        pct = dt / base * 100 if base else 0
        print(f"\n  Δ {OUR_LABEL} − {REF_LABEL}: "
              f"total {dt:+.2f} ({pct:+.1f}%), "
              f"машин {r['n_vehicles']-ref['n_vehicles']:+d}, "
              f"грузч {r['n_loaders']-ref['n_loaders']:+d}")


THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUB_FILL = PatternFill('solid', start_color='D9E1F2')
WIN_FILL = PatternFill('solid', start_color='C6EFCE')
LOSE_FILL = PatternFill('solid', start_color='FFC7CE')


def _h(cell, text, fill=HEADER_FILL, color='FFFFFF'):
    cell.value = text
    cell.font = Font(bold=True, color=color, name='Calibri')
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER


def _c(cell, val, fmt=None, bold=False, fill=None):
    cell.value = val
    cell.font = Font(bold=bold, name='Calibri')
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def export_excel(results, out_path):
    labels = list(results.keys())

    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        sheet_name = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Сравнение'

    _h(ws.cell(1, 1), 'Метрика')
    for col, label in enumerate(labels, 2):
        fill = SUB_FILL if label == REF_LABEL else HEADER_FILL
        color = '000000' if label == REF_LABEL else 'FFFFFF'
        _h(ws.cell(1, col), label, fill=fill, color=color)

    diff_col = None
    if REF_LABEL in results and OUR_LABEL in results:
        diff_col = len(labels) + 2
        _h(ws.cell(1, diff_col), f'Δ {OUR_LABEL}−{REF_LABEL}',
           fill=HEADER_FILL)

    rows = [
        ('Машин', lambda r: r['n_vehicles'],          '0',          False),
        ('Грузчиков', lambda r: r['n_loaders'],           '0',          False),
        ('Расстояние', lambda r: r['total_dist'],
         '#,##0.00',   False),
        ('Время работы гр', lambda r: r['loader_work_time'],
         '0',          False),
        (None, None, None, None),
        ('Аренда машин', lambda r: r['cost']['vehicles'],
         '#,##0.00',   False),
        ('Аренда грузч', lambda r: r['cost']['loaders'],
         '#,##0.00',   False),
        ('Топливо', lambda r: r['cost']['fuel'],
         '#,##0.00',   False),
        ('Работа грузч', lambda r: r['cost']['loader_w'],
         '#,##0.00',   False),
        ('Штраф опц', lambda r: r['cost']['penalty'],     '#,##0.00',   False),
        ('ИТОГО', lambda r: r['cost']['total'],       '#,##0.00',   True),
        (None, None, None, None),
        ('Обяз. покрытие',
         lambda r: f"{r['served_mandatory']}/{r['total_mandatory']}",
         None, False),
        ('Опц. покрытие',
         lambda r: f"{r['served_optional']}/{r['total_optional']}",
         None, False),
        ('Пропущ. опц',
         lambda r: ', '.join(map(str, r['missed_optional'])) or '-',
         None, False),
        (None, None, None, None),
        ('Заказов/машину',
         lambda r: r['avg_orders_per_vehicle'], '0.00', False),
        ('Слотов/грузч',
         lambda r: r['avg_slots_per_loader'], '0.00', False),
        ('Загрузка ТС',
         lambda r: r['avg_util'], '0.0%', False),
        ('Длины маршрутов',
         lambda r: str(r['route_lens']), None, False),
        ('Длины цепочек',
         lambda r: str(r['chain_lens']), None, False),
        (None, None, None, None),
        ('Multi-trip машин',
         lambda r: r['multi_trip_vehicles'], '0', False),
        ('Распред. рейсов/машину',
         lambda r: str(r['trips_dist']), None, False),
    ]

    row = 2
    for name_m, fn, fmt, bold in rows:
        if name_m is None:
            row += 1
            continue
        _c(ws.cell(row, 1), name_m, bold=True, fill=SUB_FILL)
        ref_val = None
        if REF_LABEL in results:
            try:
                ref_val = fn(results[REF_LABEL])
            except Exception:
                ref_val = None

        for col, label in enumerate(labels, 2):
            val = fn(results[label])
            is_num = isinstance(val, (int, float))
            highlight = None
            has_ref = isinstance(ref_val, (int, float))
            if bold and is_num and label != REF_LABEL and has_ref:
                highlight = WIN_FILL if val < ref_val else LOSE_FILL
            _c(ws.cell(row, col), val, fmt=fmt, bold=bold, fill=highlight)

        if diff_col is not None and isinstance(ref_val, (int, float)):
            val = fn(results[OUR_LABEL])
            if isinstance(val, (int, float)):
                d = val - ref_val
                fill = WIN_FILL if d < 0 else LOSE_FILL if d > 0 else None
                _c(ws.cell(row, diff_col), d, fmt=fmt, fill=fill)
        row += 1

    ws.column_dimensions['A'].width = 22
    for col in range(2, len(labels) + 2 + (1 if diff_col else 0)):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 16

    wb.save(out_path)
    print(f"\n[excel] сохранено: {out_path}")


if __name__ == '__main__':
    inp, results = analyze()
    print_results(results)
    if results:
        export_excel(results, 'comparison.xlsx')
