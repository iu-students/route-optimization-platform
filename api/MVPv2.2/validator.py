import argparse
import json
import logging
import os
from datetime import datetime

import pandas as pd
import numpy as np
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class NAMES:
    ID = 'id'
    ROUTE = 'route'
    TIME = 'time'

    LOAD_SS = 'loader_shift_size'
    VEH_SS = 'vehicle_shift_size'
    LOAD_TIME = 'load_time'
    VEH_CAPACITY = 'vehicle_capacity'
    VEH_SPEED = 'vehicle_speed'
    LOAD_SPEED = 'loader_speed'

    TIME_WDW = 'time_window'
    VOLUME = 'volume'
    LOADER_CNT = 'loader_cnt'
    VEH_ST = 'vehicle_service_time'
    LOAD_ST = 'loader_service_time'
    OPTIONAL = 'optional'
    X = 'x'
    Y = 'y'
    X1 = X + '_1'
    Y1 = Y + '_1'
    X2 = X + '_2'
    Y2 = Y + '_2'

    OPTIONAL_PENALTY = 'optional_order_penalty'
    LOAD_SALARY = 'loader_salary'
    VEH_SALARY = 'vehicle_salary'
    VEH_FC = 'fuel_cost'
    LOAD_W = 'loader_work'


class OUTPUT:
    veh_saf = 'DepotStartAndFinish'
    veh_ss = 'VehicleShiftSize'
    veh_cap = 'VehicleCapacity'
    load_saf = 'LoaderStartAndFinish'
    load_ss = 'LoaderShiftSize'
    TW_VIOL = 'TimeWindowViolation'
    unsch_orders = 'MandatoryUnscheduledOrders'
    veh_route_seq = 'VehicleRouteSequence'
    load_route_seq = 'LoaderRouteSequence'
    invalid_orders = 'InvalidOrders'

    veh_fc = 'VehicleFuelCost'
    veh_shifts = 'VehicleShifts'
    load_shifts = 'LoaderShifts'
    load_wt = 'LoaderWorkTime'
    unsch_optional = 'OptionalUnscheduledOrders'


def strict_round(nums: pd.Series | float, accuracy: int = 0) -> pd.Series | float:
    """Математическое округление для неотрицательных чисел"""
    if accuracy == 0:
        return np.floor(nums + 0.5)
    else:
        multiplier = 10 ** accuracy
        return np.floor(nums * multiplier + (0.5 * np.sign(nums))) / multiplier


RouteTypes = Literal['Vehicle', 'Loader']


class Validator:
    def __init__(self, input_file_path, result_file_path):
        self._logger = logging.getLogger("Validator")
        self._logger.setLevel(0)
        self._logger.disabled = False

        self.orders = pd.DataFrame()
        self.weights = {}
        self.depot = {}
        self.params = {}

        self._read_input_data(input_file_path)

        self.vehicle_routes = pd.DataFrame()
        self.loader_routes = pd.DataFrame()

        self._read_result_data(result_file_path)

        self.violations = {
            OUTPUT.veh_route_seq: 0,
            OUTPUT.load_route_seq: 0,
            OUTPUT.invalid_orders: 0,
            OUTPUT.veh_saf: 0,
            OUTPUT.veh_ss: 0,
            OUTPUT.veh_cap: 0,
            OUTPUT.load_saf: 0,
            OUTPUT.load_ss: 0,
            OUTPUT.TW_VIOL: 0,
            OUTPUT.unsch_orders: 0,
        }

        self.costs = {
            OUTPUT.veh_fc: [0, self.weights[NAMES.VEH_FC], 0],
            OUTPUT.veh_shifts: [0, self.weights[NAMES.VEH_SALARY], 0],
            OUTPUT.load_shifts: [0, self.weights[NAMES.LOAD_SALARY], 0],
            OUTPUT.load_wt: [0, self.weights[NAMES.LOAD_W], 0],
            OUTPUT.unsch_optional: [0, self.weights[NAMES.OPTIONAL_PENALTY], 0],
        }

    def _dist(self, point_id_1, point_id_2):
        """Расчёт расстояния между точками"""
        return strict_round(np.linalg.norm(np.array(self.orders.loc[point_id_1, [NAMES.X, NAMES.Y]])
                                           - np.array(self.orders.loc[point_id_2, [NAMES.X, NAMES.Y]])), 2)

    def _dist_to_depot(self, point_id):
        """Расчёт расстояния от точки до склада"""
        return self._dist(point_id, 0)

    def _read_input_data(self, input_file_path) -> None:
        """Чтение и обработка файла с примером"""

        with open(input_file_path, "r") as input_file:
            input_data = json.load(input_file)
        self.orders = pd.concat([
            pd.DataFrame({
                NAMES.ID: [0],
                NAMES.X: [input_data["depot"][NAMES.X]],
                NAMES.Y: [input_data["depot"][NAMES.Y]],
                NAMES.TIME_WDW: [[0, 100000]],
                NAMES.VEH_ST: [input_data["depot"][NAMES.LOAD_TIME]],
                NAMES.VOLUME: [0]
            }),
            pd.DataFrame(input_data["orders"]),
        ], ignore_index=True).set_index(NAMES.ID, drop=False)
        self.orders.index.name = 'index'

        self.weights = input_data["weights"]
        self.depot = input_data["depot"]
        self.params = {k: v for k, v in input_data.items() if k not in ["orders", "weights", "depot"]}

    def _read_result_data(self, result_file_path) -> None:
        """Чтение файла с результатом"""

        with open(result_file_path, "r") as result_file:
            result_data = json.load(result_file)

        self.vehicle_routes = pd.DataFrame(result_data["vehicles"])
        self.loader_routes = pd.DataFrame(result_data["loaders"])

    def _proc_result_data(self):
        """Обработка данных результата"""

        self.vehicle_routes = (
            self.vehicle_routes
            .assign(**{
                NAMES.TIME: lambda df: df.apply(
                    lambda row:
                    [strict_round(
                        row[NAMES.TIME][0] - strict_round(
                            self._dist_to_depot(row[NAMES.ROUTE][1]) / self.params[NAMES.VEH_SPEED], 2), 2)]
                    + row[NAMES.TIME]
                    + [strict_round(
                        row[NAMES.TIME][-1] + strict_round(
                            self._dist_to_depot(row[NAMES.ROUTE][-2]) / self.params[NAMES.VEH_SPEED], 2)
                        + self.orders.loc[row[NAMES.ROUTE][-2], NAMES.VEH_ST],
                        2)], axis=1),
            })
        )
        veh_routes_orders = self.vehicle_routes[NAMES.ROUTE].sum()
        self.loader_routes = (
            self.loader_routes
            .assign(**{
                NAMES.ROUTE: lambda df: df[NAMES.ROUTE].apply(lambda x: [_ for _ in x if _ in veh_routes_orders]),
            })
            .assign(route_len=lambda df: df[NAMES.ROUTE].apply(lambda x: len(x)))
            .query('route_len > 0')
            .drop(columns='route_len')
            .assign(**{
                NAMES.ROUTE: lambda df: df.apply(lambda row: row[NAMES.ROUTE] + [row[NAMES.ROUTE][0]], axis=1),
            })
        )

    def _correct_orders_id(self, routes: pd.DataFrame, name: RouteTypes):
        """ Проверка на правильность id заказов """

        assert name in ['Vehicle', 'Loader']

        invalid_id = (
            routes
            .explode(NAMES.ROUTE)
            .query(f"{NAMES.ROUTE} != 0")
            .assign(invalid=lambda df: ~(df[NAMES.ROUTE].isin(self.orders[NAMES.ID])))
            .query("invalid")
        )
        if not invalid_id.empty:
            self._logger.warning(f"\n{name} routes contains invalid orders:\n "
                                 + str(invalid_id[[NAMES.ID, NAMES.ROUTE]].drop_duplicates().groupby(NAMES.ID)
                                       .agg(**{"Invalid orders": pd.NamedAgg(NAMES.ROUTE, lambda x: list(x))})))
            self.violations[OUTPUT.invalid_orders] += invalid_id.shape[0]

            if name == 'Vehicle':
                self.vehicle_routes[NAMES.TIME] = self.vehicle_routes[[NAMES.TIME, NAMES.ROUTE]].apply(
                    lambda row: [row[NAMES.TIME][i] for i in range(len(row[NAMES.TIME]))
                                 if row[NAMES.ROUTE][i + 1] in self.orders.loc[:, NAMES.ID]],
                    axis=1
                )
                self.vehicle_routes[NAMES.ROUTE] = self.vehicle_routes[NAMES.ROUTE].apply(
                    lambda x: [_ for _ in x if _ in self.orders.loc[:, NAMES.ID]]
                )
            else:  # name == 'Loader'
                self.loader_routes[NAMES.ROUTE] = self.loader_routes[NAMES.ROUTE].apply(
                    lambda x: [_ for _ in x if _ in self.orders.loc[:, NAMES.ID]]
                )

    def _correct_start_time(self, routes: pd.DataFrame, name: RouteTypes):
        """Проверка на корректность времени начала работ"""

        assert name in ['Vehicle', 'Loader']

        service_time_name = NAMES.VEH_ST if name == 'Vehicle' else NAMES.LOAD_ST
        speed = self.params[NAMES.VEH_SPEED] if name == "Vehicle" else self.params[NAMES.LOAD_SPEED]

        if name == 'Vehicle':
            points = (
                routes
                .assign(
                    point=routes[NAMES.ROUTE].apply(lambda x: x[:-1]),
                    next_point=routes[NAMES.ROUTE].apply(lambda x: x[1:]),
                    time=routes[NAMES.TIME].apply(lambda x: x[:-1]),
                    next_time=routes[NAMES.TIME].apply(lambda x: x[1:])
                )
                .explode(['next_point', 'point', 'time', 'next_time'])
            )
        else:  # name == 'Loader'
            points = (
                routes
                .assign(
                    point=routes[NAMES.ROUTE].apply(lambda x: x[:-1]),
                    next_point=routes[NAMES.ROUTE].apply(lambda x: x[1:]),
                )
                .explode(['next_point', 'point'])
                .merge(self.vehicle_routes[[NAMES.ROUTE, NAMES.TIME]]
                       .explode([NAMES.ROUTE, NAMES.TIME])
                       .rename(columns={NAMES.ROUTE: 'point', NAMES.TIME: 'time'}),
                       on=['point'], how='left')
                .merge(self.vehicle_routes[[NAMES.ROUTE, NAMES.TIME]]
                       .explode([NAMES.ROUTE, NAMES.TIME])
                       .rename(columns={NAMES.ROUTE: 'next_point', NAMES.TIME: 'next_time'}),
                       on=['next_point'], how='left')
                .assign(
                    next_time=lambda df: df.apply(
                        lambda row: strict_round(
                            self.orders.loc[row['point'], NAMES.LOAD_ST] + row['time']
                            + strict_round(self._dist(row['point'], row['next_point']) / speed,
                                           2),
                            2)
                        if row[NAMES.ROUTE][-1] == row['next_point'] else row['next_time'],
                        axis=1)
                )
            )

        points = (
            points
            .merge(self.orders[[NAMES.ID, NAMES.X, NAMES.Y, service_time_name, NAMES.TIME_WDW]]
                   .rename(columns={NAMES.ID: 'point_id'}),
                   left_on='point', right_on='point_id',
                   how='left')
            .merge(self.orders[[NAMES.ID, NAMES.X, NAMES.Y]]
                   .rename(columns={NAMES.ID: 'point_id'}),
                   left_on='next_point', right_on='point_id', how='left',
                   suffixes=('_1', '_2'))
        )
        if name == 'Vehicle':
            # Находим индексы строк с минимальным временем для каждого id
            route_start_indices = points.groupby(NAMES.ID)['time'].idxmin()
            # Устанавливаем время погрузки на складе 0 для этих строк
            points.loc[route_start_indices, NAMES.VEH_ST] = 0

        points = (
            points
            .assign(
                travel_time=lambda df: strict_round(
                    strict_round(np.sqrt((df[NAMES.X1] - df[NAMES.X2]) ** 2 + (df[NAMES.Y1] - df[NAMES.Y2]) ** 2), 2)
                    / speed, 2),

                # успеваем приехать к времени начала разгрузки
                min_correct_time=lambda df: strict_round(df.time + df[service_time_name] + df.travel_time, 2),
                is_achievable=lambda df: df['min_correct_time'] <= df.next_time,
            )
        )
        if name == 'Vehicle':
            points = (
                points
                .assign(
                    # разгрузка внутри таймслота
                    in_window=lambda df: df.apply(
                        lambda row: row[NAMES.TIME_WDW][0] <= row['time'] <= row[NAMES.TIME_WDW][1], axis=1)
                )
            )
            not_in_window = points.query("not in_window")
            if not not_in_window.empty:
                self._logger.warning(f"\n{name} start times do not fall within the time windows:\n "
                                     + str(not_in_window[[NAMES.ID, 'point']].drop_duplicates().groupby(NAMES.ID)
                                           .agg(**{"Invalid orders": pd.NamedAgg('point', lambda x: list(x))}))
                                     )
                self.violations[OUTPUT.TW_VIOL] = not_in_window.shape[0]

        unachievable = points.query("not is_achievable")
        if not unachievable.empty:
            self._logger.warning(
                f"\n{name} incorrect times:\n "
                + str(unachievable
                      .assign(released=lambda df: df['time'] + df[service_time_name])
                      [[NAMES.ID, 'next_point', 'released', 'travel_time', 'min_correct_time', 'next_time']]
                      .rename(
                    columns={NAMES.ID: f'{name}_id', 'next_point': 'point',
                             'next_time': 'solution_time', })
                      .set_index(f'{name}_id')
                      .to_string()
                      )
            )
            viol_name = OUTPUT.veh_route_seq if name == 'Vehicle' else OUTPUT.load_route_seq
            self.violations[viol_name] = unachievable.shape[0]

        shift_max = self.params[(NAMES.VEH_SS if name == "Vehicle" else NAMES.LOAD_SS)]
        shift_length = (
            points
            .groupby(NAMES.ID)
            .agg({'time': 'first', 'next_time': 'last'})
            .assign(shift_time=lambda df: df['next_time'] - df['time'])
        )

        shift_length = (
            shift_length
            .query(f'shift_time >= {shift_max}')
        )
        if not shift_length.empty:
            self._logger.warning(f"\n{name} exceeding the shift duration: "
                                 + ', '.join([str(_) for _ in shift_length.index.astype(int).values])
                                 )
            self.violations[f'{name}{OUTPUT.load_ss[6:]}'] = shift_length.shape[0]

    def _correct_capacities(self, routes: pd.DataFrame):
        """Проверка, что грузоподъёмность не нарушается"""

        def route_to_circles(route: list):
            route = ['-' if i == 0 else i for i in route]
            circles = [[int(i) for i in group.split()] for group in ' '.join(map(str, route)).split('-') if group]
            return circles

        capacities = (
            routes
            .assign(
                circles=lambda df: df[NAMES.ROUTE].apply(route_to_circles),
                circles_num=lambda df: df.circles.apply(lambda x: list(range(len(x)))),
            )
            .explode(['circles', 'circles_num'])
            .explode('circles')
            .merge(self.orders[[NAMES.ID, NAMES.VOLUME]].rename(columns={NAMES.ID: 'circles'}),
                   on='circles', how='left')
            .groupby(NAMES.ID)
            .agg({NAMES.VOLUME: 'sum'})
            .query(f'{NAMES.VOLUME} > {self.params[NAMES.VEH_CAPACITY]}')
        )
        if not capacities.empty:
            self._logger.warning(f"\nVehicle with violation of capacity: "
                                 + ', '.join([str(_) for _ in capacities.index.values])
                                 )
            self.violations[OUTPUT.veh_cap] = capacities.shape[0]

    def _complete_order(self):
        """Проверка, что заявки выполнены"""
        orders = (
            self.orders
            .query(f'{NAMES.ID} != 0')
            .merge(self.vehicle_routes.explode([NAMES.ROUTE, NAMES.TIME])
                   .rename(columns={NAMES.ID: 'veh_id'}),
                   how='left', left_on=NAMES.ID, right_on=NAMES.ROUTE)
            .merge(self.loader_routes.explode([NAMES.ROUTE])
                   .rename(columns={NAMES.ID: 'loader_id'}),
                   how='left', left_on=NAMES.ID, right_on=NAMES.ROUTE)
        )

        complete = (
            orders
            .groupby([NAMES.ID, 'veh_id', NAMES.LOADER_CNT, NAMES.OPTIONAL], as_index=False, dropna=False)
            .agg({'loader_id': lambda x: sum(x.notna())})
            .assign(
                loaders_enought=lambda df: df[NAMES.LOADER_CNT] <= df['loader_id'],
                no_veh=lambda df: df['veh_id'].isna(),
                incomplete=lambda df: ~df['loaders_enought'] | df['no_veh']
            )
            .groupby([NAMES.ID, NAMES.OPTIONAL], as_index=False, dropna=False)
            .agg({"incomplete": lambda x: all(x)})
        )
        incomplete_obligatory = complete.query(f"{NAMES.OPTIONAL} == 0 and incomplete")
        if not incomplete_obligatory.empty:
            self._logger.warning(
                f"\nIncomplete obligatory orders: {', '.join(incomplete_obligatory[NAMES.ID].astype(str).values)}")
            self.violations[OUTPUT.unsch_orders] = incomplete_obligatory.shape[0]

        incomplete_optional = complete.query(f"{NAMES.OPTIONAL} == 1 and incomplete")
        if not incomplete_optional.empty:
            self._logger.warning(
                f"\nIncomplete optional orders: {', '.join(incomplete_optional[NAMES.ID].astype(str).values)}")
            self.costs[OUTPUT.unsch_optional][2] = incomplete_optional.shape[0]
            self.costs[OUTPUT.unsch_optional][0] = (self.costs[OUTPUT.unsch_optional][1]
                                                    * self.costs[OUTPUT.unsch_optional][2])

    def _start_and_finish(self):
        """Проверка начала и окончания маршрутов"""

        # Проверка, что у vehicle начало и окончание -- depot
        depot_start_and_finish = (
            self.vehicle_routes
            .assign(is_correct=lambda df: df[NAMES.ROUTE].apply(lambda x: x[0] == x[-1] == 0))
            .query('not is_correct')
        )
        if not depot_start_and_finish.empty:
            self._logger.warning(
                f"\nVehicle routes have to start and finish in depot. \n"
                f"Incorrect routes for {', '.join(depot_start_and_finish[NAMES.ID].astype(str).values)} vehicle")
            self.violations[OUTPUT.veh_saf] = depot_start_and_finish.shape[0]

        # Проверка, что у грузчиков начало и окончание совпадают
        loader_start_and_finish = (
            self.loader_routes
            .assign(is_correct=lambda df: df[NAMES.ROUTE].apply(lambda x: x[0] == x[-1]))
            .query('not is_correct')
        )
        if not loader_start_and_finish.empty:
            self._logger.warning(
                f"\nLoader routes have to start and finish in same point. \n"
                f"Incorrect routes for {', '.join(loader_start_and_finish[NAMES.ID].astype(str).values)} vehicle")
            self.violations[OUTPUT.load_saf] = loader_start_and_finish.shape[0]

    def _route_dist_calc(self, routes: pd.DataFrame, name: RouteTypes) -> float:
        """ Расчёт времени перемещений """

        assert name in ['Vehicle', 'Loader']

        speed = self.params[NAMES.VEH_SPEED] if name == "Vehicle" else self.params[NAMES.LOAD_SPEED]
        dist_times = (
            routes
            .assign(
                point=routes[NAMES.ROUTE].apply(lambda x: x[:-1]),
                next_point=routes[NAMES.ROUTE].apply(lambda x: x[1:]),
            )
            .explode(['next_point', 'point'])
            .merge(self.orders[[NAMES.ID, NAMES.X, NAMES.Y]]
                   .rename(columns={NAMES.ID: 'point_id'}),
                   left_on='point', right_on='point_id',
                   how='left')
            .merge(self.orders[[NAMES.ID, NAMES.X, NAMES.Y]].rename(columns={NAMES.ID: 'point_id'}),
                   left_on='next_point', right_on='point_id', how='left',
                   suffixes=('_1', '_2'))
            .assign(
                travel_time=lambda df: strict_round(
                    strict_round(np.sqrt((df[NAMES.X1] - df[NAMES.X2]) ** 2 + (df[NAMES.Y1] - df[NAMES.Y2]) ** 2), 2)
                    / speed, 2)
            )
        )
        return strict_round(dist_times['travel_time'].sum(), 2)

    def _route_time_calc(self, routes: pd.DataFrame, name: RouteTypes) -> float:
        """Расчёт времени маршрута"""

        assert name in ['Vehicle', 'Loader']

        times = (
            routes
            .assign(
                start_point=routes[NAMES.ROUTE].apply(lambda x: x[0]),
                end_point=routes[NAMES.ROUTE].apply(lambda x: x[-1]),
            )
        )
        if name == 'Loader':
            times = (
                times
                .merge(
                    self.vehicle_routes[[NAMES.ROUTE, NAMES.TIME]].explode([NAMES.ROUTE, NAMES.TIME])
                    .rename(columns={NAMES.TIME: 'start_time'}),
                    how='left', left_on='start_point', right_on=NAMES.ROUTE
                )
                .merge(
                    self.vehicle_routes[[NAMES.ROUTE, NAMES.TIME]].explode([NAMES.ROUTE, NAMES.TIME])
                    .rename(columns={NAMES.TIME: 'end_time'}),
                    how='left', left_on='end_point', right_on=NAMES.ROUTE
                )
                .merge(self.orders[[NAMES.ID, NAMES.LOAD_ST]].rename(columns={NAMES.ID: 'end_point'}),
                       how='left', on='end_point')
                .assign(
                    end_time=lambda df: df['end_time'] + df[NAMES.LOAD_ST],
                )
            )
        else:  # name == 'Vehicle'
            times = times.assign(
                start_time=routes[NAMES.TIME].apply(lambda x: x[0]),
                end_time=routes[NAMES.TIME].apply(lambda x: x[-1]),
            )

        return strict_round((times['end_time'] - times['start_time']).sum(), 2)

    def _output(self):
        """Вывод всех нарушений"""
        separator = '.'
        min_spacing = 3
        max_key_len = max([len(str(k)) for k in self.violations.keys()] + [len(str(k)) for k in self.costs.keys()])
        max_val_len = max(
            [len(str(k)) for k in self.violations.values()] + [len(str(k[0])) for k in self.costs.values()])

        common_len = max_key_len + min_spacing + max_val_len

        self._logger.warning('VIOLATIONS:')
        for key, value in self.violations.items():
            dots = separator.ljust(common_len - len(str(key)) - len(str(value)), '.')
            self._logger.warning(f"{key}{dots}{value}")
        self._logger.warning(f"Total violations = {sum(self.violations.values())}")

        self._logger.warning('\nCOSTS (weight X cost):')
        for key, value in self.costs.items():
            dots = separator.ljust(common_len - len(str(key)) - len(str(value[0])), '.')
            self._logger.warning(f"{key}{dots}{value[0]} ({value[1]} X {value[2]})")
        self._logger.warning(f"Total cost = {strict_round(sum([v[0] for v in self.costs.values()]), 2)}")

    def report_rows(self) -> list[tuple]:
        """Список строк (метка, значение, формат, жирный) для экспорта в Excel.
        Порядок и состав строк одинаковы для любого Validator над одним и тем же input,
        что позволяет сопоставлять строки result vs baseline построчно по индексу."""
        rows: list[tuple] = [
            ('Машин', self.vehicle_routes.shape[0], '0', False),
            ('Грузчиков', self.loader_routes.shape[0], '0', False),
            (None, None, None, None),
            ('НАРУШЕНИЯ', None, None, None),
        ]
        for key, val in self.violations.items():
            rows.append((key, val, '0', False))
        rows.append(('Всего нарушений', self.total_violations(), '0', True))
        rows.append((None, None, None, None))
        rows.append(('СТОИМОСТЬ (weight X cost)', None, None, None))
        for key, val in self.costs.items():
            rows.append((key, val[0], '#,##0.00', False))
        rows.append(('ИТОГО', self.total_cost(), '#,##0.00', True))
        return rows

    def total_cost(self) -> float:
        """Итоговая стоимость решения (сумма всех компонент costs)"""
        return strict_round(sum(v[0] for v in self.costs.values()), 2)

    def total_violations(self) -> int:
        """Суммарное кол-во нарушений жёстких ограничений"""
        return sum(self.violations.values())

    def compare_with_baseline(self, baseline: "Validator", our_label: str = "OUR",
                              baseline_label: str = "BASELINE") -> None:
        """Сравнение текущего решения с решением-бейзлайном (тот же input)"""

        self._logger.warning(f"\n{'='*60}")
        self._logger.warning(f"COMPARISON: {our_label} vs {baseline_label}")
        self._logger.warning(f"{'='*60}")

        separator = '.'
        min_spacing = 3
        keys = list(self.costs.keys())
        max_key_len = max(len(str(k)) for k in keys)

        self._logger.warning(f"\nCOSTS ({baseline_label} -> {our_label}, delta):")
        for key in keys:
            base_val = baseline.costs[key][0]
            our_val = self.costs[key][0]
            delta = round(our_val - base_val, 2)
            dots = separator.ljust(max_key_len - len(str(key)) + min_spacing, '.')
            self._logger.warning(f"{key}{dots}{base_val:>12.2f} -> {our_val:>12.2f}  ({delta:+.2f})")

        base_total = baseline.total_cost()
        our_total = self.total_cost()
        delta_total = round(our_total - base_total, 2)
        pct = (delta_total / base_total * 100) if base_total else 0.0
        self._logger.warning(f"\nTotal cost: {base_total:.2f} -> {our_total:.2f}  "
                             f"({delta_total:+.2f}, {pct:+.1f}%)")

        base_viol = baseline.total_violations()
        our_viol = self.total_violations()
        self._logger.warning(f"Total violations: {base_viol} -> {our_viol}  "
                             f"({our_viol - base_viol:+d})")

        base_veh = baseline.vehicle_routes.shape[0]
        our_veh = self.vehicle_routes.shape[0]
        base_load = baseline.loader_routes.shape[0]
        our_load = self.loader_routes.shape[0]
        self._logger.warning(f"Vehicles: {base_veh} -> {our_veh}  ({our_veh - base_veh:+d})")
        self._logger.warning(f"Loaders:  {base_load} -> {our_load}  ({our_load - base_load:+d})")

        if our_viol > 0:
            self._logger.warning(
                "\n!!! Внимание: решение нарушает жёсткие ограничения, "
                "сравнение по стоимости может быть некорректным.")

    def validation(self, verbose: bool = True):
        """Валидация результата"""

        # Проверка на уникальность id ТС
        if not self.vehicle_routes[NAMES.ID].is_unique:
            self._logger.warning("Vehicle IDs is not unique. The IDs will be replaced with consecutive ones.")
            self.vehicle_routes[NAMES.ID] = np.arange(1, self.vehicle_routes.shape[0] + 1)

        # Проверка на уникальность id заказов в маршрутах
        if not self.vehicle_routes.explode(NAMES.ROUTE).query(f"{NAMES.ROUTE}!=0")[NAMES.ROUTE].is_unique:
            raise Exception("Orders in vehicle routes is not unique")

        # Проверка на уникальность id грузчиков
        if not self.loader_routes[NAMES.ID].is_unique:
            self._logger.warning("Loaders IDs is not unique. The IDs will be replaced with consecutive ones.")
            self.loader_routes[NAMES.ID] = np.arange(1, self.loader_routes.shape[0] + 1)

        # Проверка на правильность id заказов
        self._correct_orders_id(self.vehicle_routes, name="Vehicle")
        self._correct_orders_id(self.loader_routes, name="Loader")

        # Добавление времён первого и последнего прибытия на склад
        self._proc_result_data()

        # Проверка начала и окончания маршрута
        self._start_and_finish()

        # Проверка на корректность времени начала работ
        self._correct_start_time(self.vehicle_routes, name="Vehicle")
        self._correct_start_time(self.loader_routes, name="Loader")

        # Проверка грузоподъёмности
        self._correct_capacities(self.vehicle_routes)

        # Проверка выполнения заявок
        self._complete_order()

        # кол-во ТС
        self.costs[OUTPUT.veh_shifts][2] = self.vehicle_routes.shape[0]
        self.costs[OUTPUT.veh_shifts][0] = (self.costs[OUTPUT.veh_shifts][1] * self.costs[OUTPUT.veh_shifts][2])
        # кол-во грузчиков
        self.costs[OUTPUT.load_shifts][2] = self.loader_routes.shape[0]
        self.costs[OUTPUT.load_shifts][0] = (self.costs[OUTPUT.load_shifts][1] * self.costs[OUTPUT.load_shifts][2])

        # Расчёт расстояний
        self.costs[OUTPUT.veh_fc][2] = self._route_dist_calc(self.vehicle_routes, name="Vehicle")
        self.costs[OUTPUT.veh_fc][0] = strict_round(self.costs[OUTPUT.veh_fc][1] * self.costs[OUTPUT.veh_fc][2], 2)
        self.costs[OUTPUT.load_wt][2] = self._route_time_calc(self.loader_routes, name="Loader")
        self.costs[OUTPUT.load_wt][0] = strict_round(self.costs[OUTPUT.load_wt][1] * self.costs[OUTPUT.load_wt][2], 2)

        if verbose:
            self._output()


THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUB_FILL = PatternFill('solid', start_color='D9E1F2')
WIN_FILL = PatternFill('solid', start_color='C6EFCE')
LOSE_FILL = PatternFill('solid', start_color='FFC7CE')


def _h(cell, text, fill=HEADER_FILL, color='FFFFFF'):
    cell.value = text
    cell.font = Font(bold=True, color=color, name='Calibri')
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER


def _c(cell, val, fmt=None, bold=False, fill=None):
    cell.value = val
    cell.font = Font(bold=bold, name='Calibri')
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def _unique_sheet_name(wb: Workbook, desired: str) -> str:
    """Excel-имя листа: максимум 31 символ, без дублей."""
    desired = desired[:31]
    if desired not in wb.sheetnames:
        return desired
    i = 1
    while True:
        suffix = f" ({i})"
        candidate = desired[:31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def export_excel(out_path: str, sheet_name: str, result: "Validator", result_label: str,
                  baseline: "Validator" = None, baseline_label: str = "BASELINE") -> str:
    """Записывает результат валидации (и, если есть, сравнение с baseline) на новый лист
    файла out_path. Если файл уже существует -- добавляет лист, не трогая остальные."""

    if os.path.exists(out_path):
        wb = load_workbook(out_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = _unique_sheet_name(wb, sheet_name)
    ws = wb.create_sheet(sheet_name)

    has_baseline = baseline is not None
    labels = ([baseline_label, result_label] if has_baseline else [result_label])

    _h(ws.cell(1, 1), 'Метрика')
    for col, label in enumerate(labels, 2):
        is_baseline_col = has_baseline and label == baseline_label
        _h(ws.cell(1, col), label,
           fill=SUB_FILL if is_baseline_col else HEADER_FILL,
           color='000000' if is_baseline_col else 'FFFFFF')

    diff_col = None
    if has_baseline:
        diff_col = len(labels) + 2
        _h(ws.cell(1, diff_col), f'\u0394 {result_label}\u2212{baseline_label}', fill=HEADER_FILL)

    result_rows = result.report_rows()
    baseline_rows = baseline.report_rows() if has_baseline else None

    row_idx = 2
    for i, (label, val, fmt, bold) in enumerate(result_rows):
        if label is None:
            row_idx += 1
            continue

        base_val = baseline_rows[i][1] if has_baseline else None

        if val is None:
            # заголовок секции (НАРУШЕНИЯ / СТОИМОСТЬ)
            _c(ws.cell(row_idx, 1), label, bold=True, fill=SUB_FILL)
            row_idx += 1
            continue

        _c(ws.cell(row_idx, 1), label, bold=bold, fill=SUB_FILL if bold else None)

        values = [base_val, val] if has_baseline else [val]
        for col_offset, v in enumerate(values):
            col = 2 + col_offset
            is_result_col = (col_offset == len(values) - 1)
            highlight = None
            if (bold and is_result_col and has_baseline
                    and isinstance(v, (int, float)) and isinstance(base_val, (int, float))):
                highlight = WIN_FILL if v < base_val else (LOSE_FILL if v > base_val else None)
            _c(ws.cell(row_idx, col), v, fmt=fmt, bold=bold, fill=highlight)

        if diff_col is not None and isinstance(val, (int, float)) and isinstance(base_val, (int, float)):
            d = round(val - base_val, 2)
            fill = WIN_FILL if d < 0 else (LOSE_FILL if d > 0 else None)
            _c(ws.cell(row_idx, diff_col), d, fmt=fmt, fill=fill)

        row_idx += 1

    ws.column_dimensions['A'].width = 28
    last_col = len(labels) + 1 + (1 if diff_col else 0)
    for col in range(2, last_col + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

    wb.save(out_path)
    return sheet_name


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", default='../data', help="Validate data in dir (default 'data')")
    parser.add_argument("--input_file", type=str, default='input',
                        help="Name of instance input file (default 'input')")
    parser.add_argument("--result_file", type=str, default='result',
                        help="Name of instance result file (default 'result')")
    parser.add_argument("--baseline_file", type=str, default=None,
                        help="Name of baseline result file to compare against (optional, e.g. 'baseline')")
    parser.add_argument("--task_name", type=str, default=None,
                        help="Task/instance name for the Excel sheet name, e.g. 'i4' "
                             "(default: same as --input_file)")
    parser.add_argument("--excel_out", type=str, default='validation_results.xlsx',
                        help="Excel file to append the result to (default 'validation_results.xlsx')")

    args = parser.parse_args()

    directory = args.dir
    input_file_name = args.input_file
    result_file_name = args.result_file
    baseline_file_name = args.baseline_file
    task_name = args.task_name or input_file_name

    input_file_path = f"{directory}/{input_file_name}.json"
    result_file_path = f"{directory}/{result_file_name}.json"

    result_validator = Validator(input_file_path, result_file_path)
    result_validator.validation()

    baseline_validator = None
    if baseline_file_name:
        baseline_file_path = f"{directory}/{baseline_file_name}.json"
        try:
            baseline_validator = Validator(input_file_path, baseline_file_path)
            baseline_validator.validation(verbose=False)
        except FileNotFoundError:
            logging.getLogger("Validator").warning(
                f"\nBaseline file not found: {baseline_file_path}. Skipping comparison.")
            baseline_validator = None
        else:
            result_validator.compare_with_baseline(baseline_validator,
                                                    our_label=result_file_name,
                                                    baseline_label=baseline_file_name)

    sheet_name = f"{datetime.now():%Y-%m-%d %H-%M-%S} {task_name}"
    saved_sheet = export_excel(args.excel_out, sheet_name, result_validator,
                               result_label=result_file_name,
                               baseline=baseline_validator,
                               baseline_label=baseline_file_name or "BASELINE")
    logging.getLogger("Validator").warning(f"\n[excel] сохранено: {args.excel_out} (лист: {saved_sheet})")