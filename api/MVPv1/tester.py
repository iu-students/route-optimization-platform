import json
import math
import os
import sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


SOLUTIONS = {
    'НАШЕ':     'vb_sol_{}.json',
    'КОЛЛЕГА':  'va_sol_{}.json',
    'BASELINE': 'sol_{}.json',
}
REF_LABEL = 'BASELINE'


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_coords(point_id, depot, orders_by_id):
    """point_id == 0 -> depot, otherwise order id"""
    if point_id == 0:
        return depot['x'], depot['y']
    order = orders_by_id[point_id]
    return order['x'], order['y']


def euclidean(p1, p2):
    return math.hypot(p1[0] - p2[0], p1[1] - p2[1])


def calc_cost(input_data, output_data):
    w = input_data['weights']
    depot = input_data['depot']
    orders = input_data['orders']
    by_id = {o['id']: o for o in orders}
    vehicles = output_data['vehicles']
    loaders = output_data.get('loaders', [])

    n_v, n_l = len(vehicles), len(loaders)

    total_dist = 0.0
    for v in vehicles:
        r = v['route']
        for a, b in zip(r[:-1], r[1:]):
            total_dist += euclidean(get_coords(a, depot, by_id),
                                    get_coords(b, depot, by_id))

    loader_work_time = sum(by_id[oid]['loader_service_time']
                           for ld in loaders for oid in ld['route'])

    visited = {pid for v in vehicles for pid in v['route'] if pid != 0}
    mand = {o['id'] for o in orders if not o.get('optional', 0)}
    opt = {o['id'] for o in orders if o.get('optional', 0)}
    missed_opt = sorted(opt - visited)
    missed_mand = sorted(mand - visited)

    cost = {
        'vehicles':  n_v * w['vehicle_salary'],
        'loaders':   n_l * w['loader_salary'],
        'fuel':      total_dist * w['fuel_cost'],
        'loader_w':  loader_work_time * w['loader_work'],
        'penalty':   len(missed_opt) * w['optional_order_penalty'],
    }
    cost['total'] = sum(cost.values())

    route_lens = Counter(len(v['route']) - 2 for v in vehicles)
    chain_lens = Counter(len(ld['route']) for ld in loaders)

    cap = input_data['vehicle_capacity']
    util = [sum(by_id[oid]['volume'] for oid in v['route'][1:-1]) / cap
            for v in vehicles] if vehicles else [0]

    return {
        'n_vehicles': n_v,
        'n_loaders': n_l,
        'total_dist': round(total_dist, 2),
        'loader_work_time': loader_work_time,
        'cost': cost,
        'missed_optional': missed_opt,
        'missed_mandatory': missed_mand,
        'served_mandatory': len(mand & visited),
        'total_mandatory': len(mand),
        'served_optional': len(opt & visited),
        'total_optional': len(opt),
        'route_lens': dict(sorted(route_lens.items())),
        'chain_lens': dict(sorted(chain_lens.items())),
        'avg_util': round(sum(util) / len(util), 4),
        'avg_orders_per_vehicle': round(
            sum(len(v['route']) - 2 for v in vehicles) / max(n_v, 1), 2),
        'avg_slots_per_loader': round(
            sum(len(ld['route']) for ld in loaders) / max(n_l, 1), 2),
    }


def analyze_scenario(name, dir_path='test_cases'):
    inp = os.path.join(dir_path, f'{name}.json')
    if not os.path.exists(inp):
        return None, {}
    input_data = load_json(inp)
    results = {}
    for label, pattern in SOLUTIONS.items():
        path = os.path.join(dir_path, pattern.format(name))
        if os.path.exists(path):
            results[label] = calc_cost(input_data, load_json(path))
    return input_data, results


def print_scenario(name, results):
    print(f"\n{'='*70}\nСЦЕНАРИЙ {name.upper()}\n{'='*70}")
    if not results:
        print("  нет решений")
        return

    for label, r in results.items():
        c = r['cost']
        print(f"\n--- {label} ---")
        print(f"  машины:   {r['n_vehicles']:>3}    {c['vehicles']:>10.2f}")
        print(f"  грузчики: {r['n_loaders']:>3}    {c['loaders']:>10.2f}")
        print(f"  топливо (dist={r['total_dist']}): {c['fuel']:>10.2f}")
        print(f"  работа грузч (t={r['loader_work_time']}): "
              f"{c['loader_w']:>10.2f}")
        print(f"  штраф опц ({len(r['missed_optional'])} шт): "
              f"{c['penalty']:>10.2f}")
        print(f"  ИТОГО:    {c['total']:>10.2f}")
        print(f"  покрытие: обяз "
              f"{r['served_mandatory']}/{r['total_mandatory']}, "
              f"опц {r['served_optional']}/{r['total_optional']}")
        if r['missed_mandatory']:
            print(f"  !!! ПРОПУЩЕНЫ ОБЯЗАТЕЛЬНЫЕ: {r['missed_mandatory']}")
        print(f"  длины маршрутов: {r['route_lens']}")
        print(f"  длины цепочек:   {r['chain_lens']}")

    if REF_LABEL in results:
        ref = results[REF_LABEL]
        for label, r in results.items():
            if label == REF_LABEL:
                continue
            dt = r['cost']['total'] - ref['cost']['total']
            base = ref['cost']['total']
            pct = dt / base * 100 if base else 0
            print(f"\n  Δ {label} − {REF_LABEL}: "
                  f"total {dt:+.2f} ({pct:+.1f}%), "
                  f"машин {r['n_vehicles']-ref['n_vehicles']:+d}, "
                  f"грузч {r['n_loaders']-ref['n_loaders']:+d}")


THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
SUB_FILL = PatternFill('solid', start_color='D9E1F2')
WIN_FILL = PatternFill('solid', start_color='C6EFCE')
LOSE_FILL = PatternFill('solid', start_color='FFC7CE')


def _h(cell, text, fill=HEADER_FILL, color='FFFFFF'):
    cell.value = text
    cell.font = Font(bold=True, color=color, name='Calibri')
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER


def _c(cell, val, fmt=None, bold=False, fill=None):
    cell.value = val
    cell.font = Font(bold=bold, name='Calibri')
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def _write_summary(ws, all_data):
    headers = ['Сценарий', 'Решение', 'Машин', 'Грузчиков',
               'Топливо', 'Работа гр.', 'Штраф', 'ИТОГО',
               'Обяз', 'Опц', 'Пропущ. опц']
    for col, h in enumerate(headers, 1):
        _h(ws.cell(1, col), h)

    row = 2
    for name, (_, results) in all_data.items():
        ref_total = (
            results[REF_LABEL]['cost']['total']
            if REF_LABEL in results else None)
        for label, r in results.items():
            c = r['cost']
            is_ref = (label == REF_LABEL)
            fill = SUB_FILL if is_ref else None

            _c(ws.cell(row, 1), name.upper(), bold=True, fill=fill)
            _c(ws.cell(row, 2), label, bold=is_ref, fill=fill)
            _c(ws.cell(row, 3), r['n_vehicles'], fill=fill)
            _c(ws.cell(row, 4), r['n_loaders'], fill=fill)
            _c(ws.cell(row, 5), c['fuel'], fmt='#,##0.00', fill=fill)
            _c(ws.cell(row, 6), c['loader_w'], fmt='#,##0.00', fill=fill)
            _c(ws.cell(row, 7), c['penalty'], fmt='#,##0.00', fill=fill)

            total_fill = fill
            if not is_ref and ref_total is not None:
                total_fill = WIN_FILL if c['total'] < ref_total else LOSE_FILL
            _c(ws.cell(row, 8), c['total'], fmt='#,##0.00',
               bold=True, fill=total_fill)

            mand = f"{r['served_mandatory']}/{r['total_mandatory']}"
            opt = f"{r['served_optional']}/{r['total_optional']}"
            missed = ', '.join(map(str, r['missed_optional'])) or '—'
            _c(ws.cell(row, 9), mand, fill=fill)
            _c(ws.cell(row, 10), opt, fill=fill)
            _c(ws.cell(row, 11), missed, fill=fill)
            row += 1
        row += 1

    for col, w in enumerate([10, 12, 8, 10, 12, 12, 12, 14, 8, 8, 22], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w


def _write_scenario(ws, name, results):
    labels = list(results.keys())

    _h(ws.cell(1, 1), f'Метрика — {name.upper()}')
    for col, label in enumerate(labels, 2):
        fill = SUB_FILL if label == REF_LABEL else HEADER_FILL
        color = '000000' if label == REF_LABEL else 'FFFFFF'
        _h(ws.cell(1, col), label, fill=fill, color=color)

    diff_cols = {}
    if REF_LABEL in results:
        next_col = len(labels) + 2
        for label in labels:
            if label == REF_LABEL:
                continue
            _h(ws.cell(1, next_col),
               f'Δ {label}−{REF_LABEL}', fill=HEADER_FILL)
            diff_cols[label] = next_col
            next_col += 1

    rows = [
        ('Машин', lambda r: r['n_vehicles'],          '0',          False),
        ('Грузчиков', lambda r: r['n_loaders'],           '0',          False),
        ('Расстояние', lambda r: r['total_dist'],
         '#,##0.00',   False),
        ('Время работы гр', lambda r: r['loader_work_time'],
         '0',          False),
        (None, None, None, None),
        ('Аренда машин', lambda r: r['cost']['vehicles'],
         '#,##0.00',   False),
        ('Аренда грузч', lambda r: r['cost']['loaders'],
         '#,##0.00',   False),
        ('Топливо', lambda r: r['cost']['fuel'],
         '#,##0.00',   False),
        ('Работа грузч', lambda r: r['cost']['loader_w'],
         '#,##0.00',   False),
        ('Штраф опц', lambda r: r['cost']['penalty'],     '#,##0.00',   False),
        ('ИТОГО', lambda r: r['cost']['total'],       '#,##0.00',   True),
        (None, None, None, None),
        ('Обяз. покрытие',
         lambda r: f"{r['served_mandatory']}/{r['total_mandatory']}",
         None, False),
        ('Опц. покрытие',
         lambda r: f"{r['served_optional']}/{r['total_optional']}",
         None, False),
        ('Пропущ. опц',
         lambda r: ', '.join(map(str, r['missed_optional'])) or '—',
         None, False),
        (None, None, None, None),
        ('Заказов/машину',
         lambda r: r['avg_orders_per_vehicle'], '0.00', False),
        ('Слотов/грузч',
         lambda r: r['avg_slots_per_loader'], '0.00', False),
        ('Загрузка ТС',
         lambda r: r['avg_util'], '0.0%', False),
        ('Длины маршрутов',
         lambda r: str(r['route_lens']), None, False),
        ('Длины цепочек',
         lambda r: str(r['chain_lens']), None, False),
    ]

    row = 2
    for name_m, fn, fmt, bold in rows:
        if name_m is None:
            row += 1
            continue
        _c(ws.cell(row, 1), name_m, bold=True, fill=SUB_FILL)
        ref_val = None
        if REF_LABEL in results:
            try:
                ref_val = fn(results[REF_LABEL])
            except Exception:
                ref_val = None

        for col, label in enumerate(labels, 2):
            val = fn(results[label])
            is_num = isinstance(val, (int, float))
            highlight = None
            has_ref = isinstance(ref_val, (int, float))
            if bold and is_num and label != REF_LABEL and has_ref:
                highlight = WIN_FILL if val < ref_val else LOSE_FILL
            _c(ws.cell(row, col), val, fmt=fmt, bold=bold, fill=highlight)

        if REF_LABEL in results and isinstance(ref_val, (int, float)):
            for label, col in diff_cols.items():
                val = fn(results[label])
                if isinstance(val, (int, float)):
                    d = val - ref_val
                    fill = (WIN_FILL if d < 0
                            else LOSE_FILL if d > 0
                            else None)
                    _c(ws.cell(row, col), d, fmt=fmt, fill=fill)
        row += 1

    ws.column_dimensions['A'].width = 22
    for col in range(2, len(labels) + 2 + len(diff_cols)):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 16


def export_excel(all_data, out_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Сводка'
    _write_summary(ws_summary, all_data)

    for name, (_, results) in all_data.items():
        if not results:
            continue
        ws = wb.create_sheet(name.upper())
        _write_scenario(ws, name, results)

    wb.save(out_path)
    print(f"\n[excel] сохранено: {out_path}")


if __name__ == '__main__':
    names = sys.argv[1:] if len(sys.argv) >= 2 else ['t1', 't2', 't3']

    all_data = {}
    for n in names:
        inp, res = analyze_scenario(n)
        if inp is None:
            print(f"[пропуск] {n}: нет входного файла")
            continue
        all_data[n] = (inp, res)
        print_scenario(n, res)

    if all_data:
        export_excel(all_data, 'comparison.xlsx')
